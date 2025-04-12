import math
import numpy as np
import pandas as pd
# import tkinter.font as tkfont
import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk  # 用于表格显示
# from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment
# from openpyxl.styles import PatternFill

# WGS84椭球体常量
a = 6378137.0  # 赤道半径，单位：米
f = 1 / 298.257223563  # 扁率
b = a * (1 - f)  # 极半径
e2 = f * (2 - f)  # 第一偏心率的平方

def lat_lon_to_ecef(lat, lon, h=0.0, a=6378137.0, f=1 / 298.257223563):
    """
    将纬度和经度转换为ECEF坐标，允许传入自定义的椭球参数。
    :param lat: 纬度
    :param lon: 经度
    :param h: 高程，默认为0
    :param a: 长半径
    :param f: 扁率
    :return: ECEF坐标 (x, y, z)
    """
    # 将纬度和经度转换为弧度
    lat_rad = math.radians(lat)
    lon_rad = math.radians(lon)

    # 计算经纬度所在的卯酉圈曲率半径
    e2 = f * (2 - f)  # 第一偏心率的平方
    N = a / math.sqrt(1 - e2 * math.sin(lat_rad) ** 2)

    # 计算ECEF坐标
    x = (N + h) * math.cos(lat_rad) * math.cos(lon_rad)
    y = (N + h) * math.cos(lat_rad) * math.sin(lon_rad)
    z = (N * (1 - f) ** 2 + h) * math.sin(lat_rad)

    return x, y, z

def ecef_to_lat_lon_ecef(x, y, z, a=6378137.0, f=1 / 298.257223563):
    """
    将ECEF坐标转换为经纬度和高程，允许传入自定义的椭球参数。
    :param x: ECEF x坐标
    :param y: ECEF y坐标
    :param z: ECEF z坐标
    :param a: 长半径
    :param f: 扁率
    :return: (纬度, 经度, 高程)
    """
    e2 = f * (2 - f)

    # 计算经度
    lon = math.atan2(y, x)  # 经度（弧度）

    # 计算P
    p = math.sqrt(x ** 2 + y ** 2)

    # 计算初始纬度
    lat = math.atan2(z, (1 - f) * p)  # 初始纬度（弧度）

    # 迭代计算精确的纬度
    for _ in range(100):
        N = a / math.sqrt(1 - e2 * math.sin(lat) ** 2)
        h = p / math.cos(lat) - N
        lat = math.atan2(z + e2 * N * math.sin(lat), p)

    # 转换为度
    lat = math.degrees(lat)
    lon = math.degrees(lon)

    return lat, lon, h

def read_file(file_path):
    """
    从TXT文件读取经纬度和高程信息。
    文件格式: 序号, 台站名称, L1(经度), B1(纬度), H(高程, 可选)
    :param file_path: 文件路径
    :return: 站点信息列表
    """
    stations = []
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            for line in file:
                if not line.strip():  # 跳过空行
                    continue
                parts = line.strip().split(',')
                if len(parts) < 4:
                    continue
                num = parts[0]
                name = parts[1]
                lon = float(parts[2])
                lat = float(parts[3])
                h = float(parts[4]) if len(parts) > 4 and parts[4].strip() else 0.0  # 如果没有高程，默认为0
                stations.append({"num": num, "name": name, "lat": lat, "lon": lon, "h": h})
    except Exception as e:
        # 异常处理，显示错误信息
        messagebox.showerror("文件读取错误", f"无法读取文件: {e}")
    return stations

def ext_B(matrix):
    n = matrix.shape[0]
    result = []
    for i in range(n):
        X, Y, Z = matrix[i]
        row = [[1, 0, 0, X, 0, -Z, Y],
               [0, 1, 0, Y, Z, 0, -X],
               [0, 0, 1, Z, -Y, X, 0]]
        result.append(row)
    return np.vstack(result)

def ext_L(matrix):
    transposed_rows = []
    for row in matrix:
        transposed_rows.append(row.reshape(-1, 1))
    stacked_matrix = np.vstack(transposed_rows)
    return stacked_matrix

#七参数计算
def compute_seven_parameters(coords1,coords2):
    Matrix1 = np.array(coords1)
    Matrix2 = np.array(coords2)
    B = -ext_B(Matrix1)
    L = ext_L(Matrix2)
    P = np.eye(B.shape[0])
    deltaX = -np.linalg.inv(B.T @ P @ B) @ B.T @ P @ L
    dx = deltaX.reshape(-1)[0]
    dy = deltaX.reshape(-1)[1]
    dz = deltaX.reshape(-1)[2]
    s = deltaX.reshape(-1)[3]
    rx = deltaX.reshape(-1)[4] / deltaX.reshape(-1)[3]
    ry = deltaX.reshape(-1)[5] / deltaX.reshape(-1)[3]
    rz = deltaX.reshape(-1)[6] / deltaX.reshape(-1)[3]
    return dx, dy, dz, rx, ry, rz, s

def compute_seven_parameters_sec(coords1,coords2):
    Matrix1 = np.array(coords1)
    Matrix2 = np.array(coords2)
    B = -ext_B(Matrix1)
    L = ext_L(Matrix2)
    P = np.eye(B.shape[0])
    deltaX = -np.linalg.inv(B.T @ P @ B) @ B.T @ P @ L
    dx = deltaX.reshape(-1)[0]
    dy = deltaX.reshape(-1)[1]
    dz = deltaX.reshape(-1)[2]
    s = deltaX.reshape(-1)[3]
    rx = deltaX.reshape(-1)[4] / deltaX.reshape(-1)[3]
    ry = deltaX.reshape(-1)[5] / deltaX.reshape(-1)[3]
    rz = deltaX.reshape(-1)[6] / deltaX.reshape(-1)[3]
    m = (s - 1) * (10**6)
    rx_sec = rx * 648000.0 / math.pi
    ry_sec = ry * 648000.0 / math.pi
    rz_sec = rz * 648000.0 / math.pi
    V = np.dot(B, deltaX) + L
    xx = math.sqrt(np.sum(V.T @ V) / (3 * Matrix1.shape[0] - 7))
    return dx, dy, dz, rx_sec, ry_sec, rz_sec, m, xx

def transform_to_second_coordinate_system(coords, params):
    """
    将坐标从第一个坐标系转换到第二个坐标系。
    :param coords: 第一个坐标系的坐标列表 (x, y, z)
    :param params: 七参数 (dx, dy, dz, rx, ry, rz, s)
    :return: 转换后的坐标列表
    """

    transformed = []
    dx, dy, dz, rx, ry, rz, s = params

    # 构建旋转矩阵
    for x, y, z in coords:
        R = np.array([
            [np.cos(ry) * np.cos(rz), np.cos(ry) * np.sin(rz), -np.sin(ry)],
            [-np.cos(rx) * np.sin(rz)+np.sin(rx)*np.sin(ry)*np.cos(rz), np.cos(rx) * np.cos(rz) + np.sin(rx) * np.sin(ry) * np.sin(rz),
             np.sin(rx)*np.cos(ry)],
            [np.sin(rx) * np.sin(rz)+np.cos(rx)*np.sin(ry)*np.cos(rz), -np.sin(rx) * np.cos(rz) + np.cos(rx) * np.sin(ry) * np.sin(rz),
             np.cos(rx)*np.cos(ry)]
        ])

        # 应用七参数转换
        new_coords = s * R @ np.array([x, y, z]) + np.array([dx, dy, dz])
        transformed.append(new_coords)

    return transformed

def compute_correction_with_seven_params(XYZ1, XYZ2, XYZ_non_public, params):
    """
    计算非公共点改正数，使用七参数将XYZ1转换为坐标系2后再计算改正数
    :param XYZ1: 公共点坐标系1的坐标
    :param XYZ2: 公共点坐标系2的坐标
    :param XYZ_non_public: 非公共点坐标
    :param params: 七参数 (dx, dy, dz, rx, ry, rz, s)
    :return: 非公共点的改正数矩阵
    """
    # 使用七参数将XYZ1转换为坐标系2下的坐标
    transformed_XYZ1 = np.array(transform_to_second_coordinate_system(XYZ1, params))

    # 步骤1：计算改正数矩阵V = XYZ2 - (转换后的XYZ1)
    V = XYZ2 - transformed_XYZ1  # n行3列矩阵

    # 获取公共点数量n和非公共点数量t
    n = XYZ1.shape[0]
    t = XYZ_non_public.shape[0]

    # 初始化用于存储距离矩阵S和权重矩阵P
    S = np.zeros((t, n))
    P = np.zeros((t, n))

    # 步骤2：计算非公共点与所有公共点之间的距离
    for i in range(t):
        for j in range(n):
            # 计算距离S[i, j] = sqrt((X_i - X_j)^2 + (Y_i - Y_j)^2 + (Z_i - Z_j)^2)
            S[i, j] = np.linalg.norm(XYZ_non_public[i] - XYZ1[j])

    # 步骤3：根据距离矩阵S计算权重矩阵P
    # P = 1 / S^2，如果距离为零，则权重设为无穷大（表示此非公共点与公共点重合）
    with np.errstate(divide='ignore', invalid='ignore'):  # 忽略除以零的警告
        P = 1 / (S**2)
        P[np.isinf(P)] = 0  # 避免无穷大的权重，设为0

    # 步骤4：计算非公共点的改正数Vf
    Vf = np.zeros((t, 3))  # 初始化非公共点改正数矩阵，大小为t行3列

    for i in range(t):
        # 权重的总和
        P_sum = np.sum(P[i])
        if P_sum != 0:
            # 计算非公共点的改正数Vf[i] = (P[i] * V的每一列) / P的总和
            Vf[i] = np.dot(P[i], V) / P_sum

    return Vf

# Tkinter GUI部分
class CoordinateConverterApp:

    def export_to_excel(self):
        # 获取界面上所有数据
        params_data = [self.param_tree.item(item)["values"] for item in self.param_tree.get_children()]
        coords1_data = [self.tree1.item(item)["values"] for item in self.tree1.get_children()]
        coords2_data = [self.tree2.item(item)["values"] for item in self.tree2.get_children()]
        diff_data = [
            [item[0], item[1], round(float(item[2]), 6), round(float(item[3]), 6), round(float(item[4]), 6),
             round(float(item[5]), 6), round(float(item[6]), 6), round(float(item[7]), 6),
             ]
            for item in [self.diff_tree.item(item)["values"] for item in self.diff_tree.get_children()]
        ]
        diff_sort_data=[[round(float(item[0]), 6), round(float(item[1]), 6), round(float(item[2]), 6), round(float(item[3]), 6), round(float(item[4]), 6),
             round(float(item[5]), 6), round(float(item[6]), 6), round(float(item[7]), 6),
             round(float(item[8]), 6), round(float(item[9]), 6), round(float(item[10]), 6), round(float(item[11]), 6),]
            for item in
            [self.diff_sort_tree.item(item)["values"] for item in self.diff_sort_tree.get_children()]
        ]

        # 获取 Mp, Mx, My, Mz 表格数据
        residual_summary_data = [
            [round(float(item[0]), 6), round(float(item[1]), 6), round(float(item[2]), 6), round(float(item[3]), 6), round(float(item[4]), 6), round(float(item[5]), 6), round(float(item[6]), 6), round(float(item[7]), 6)]
            for item in
            [self.residual_summary_tree.item(item)["values"] for item in self.residual_summary_tree.get_children()]
        ]

        coords_to_transform_data = [self.transform_input_tree.item(item)["values"] for item in
                                    self.transform_input_tree.get_children()]
        transformed_data = [self.transformed_output_tree.item(item)["values"] for item in
                            self.transformed_output_tree.get_children()]

        # 新增：获取改正数表格数据
        correction_data = [
            self.correction_tree.item(item)["values"]
            for item in self.correction_tree.get_children()
        ]

        # 新增：获取改正后数据表格数据
        corrected_data = [
            self.corrected_tree.item(item)["values"]
            for item in self.corrected_tree.get_children()
        ]

        if not params_data or not coords1_data or not coords2_data or not transformed_data or not coords_to_transform_data:
            messagebox.showerror("错误", "没有可导出的数据")
            return

        # 数据字典，用于将不同表格的数据写入不同的Excel Sheet
        data_dict = {
            "七参数": pd.DataFrame(params_data, columns=["dx(m)", "dy(m)", "dz(m)", "rx(s)", "ry(s)", "rz(s)", "m(ppm)" ,"单位权中误差"]),
            "原坐标": pd.DataFrame(coords1_data,
                                   columns=["序号", "台站名称", "纬度", "经度", "高程", "ECEF X", "ECEF Y", "ECEF Z"]),
            "新坐标": pd.DataFrame(coords2_data,
                                   columns=["序号", "台站名称", "纬度", "经度", "高程", "ECEF X", "ECEF Y", "ECEF Z"]),
            "残差": pd.DataFrame(diff_data,
                             columns=["序号", "台站名称", "纬度残差", "经度残差", "高程残差", "X残差", "Y残差", "Z残差",]),
            "残差最值": pd.DataFrame(diff_sort_data,
                                 columns=["最大纬度残差", "最小纬度残差", "最大经度残差", "最小经度残差", "最大高程残差", "最小高程残差",
                                          "最大X残差", "最小X残差", "最大Y残差", "最小Y残差", "最大Z残差", "最小Z残差"]),
            "点位中误差": pd.DataFrame(residual_summary_data, columns=["Mx", "My", "Mz", "Mp", "3Mx", "3My", "3Mz", "3Mp"]),  # 新增的表格数据
            "转换前坐标": pd.DataFrame(coords_to_transform_data,
                                       columns=["序号", "台站名称", "纬度", "经度", "高程", "ECEF X", "ECEF Y",
                                                "ECEF Z"]),
            "转换后坐标": pd.DataFrame(transformed_data,
                                       columns=["序号", "台站名称", "纬度", "经度", "高程", "ECEF X", "ECEF Y",
                                                "ECEF Z"]),
            # 新增：改正数和改正后数据
            "改正数": pd.DataFrame(correction_data, columns=["序号", "台站名称", "Vx", "Vy", "Vz"]),
            "改正后坐标": pd.DataFrame(corrected_data,
                                       columns=["序号", "台站名称", "纬度(改)", "经度(改)", "高程(改)", "X(改)",
                                                "Y(改)", "Z(改)"])
        }

        # 保存为Excel文件
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            # 创建汇总表
            summary_data = []

            # 添加七参数到汇总表
            summary_data.append(["七参数"])
            summary_data.append(["dx(m)", "dy(m)", "dz(m)", "rx(s)", "ry(s)", "rz(s)", "m(ppm)" ,"单位权中误差"])
            summary_data.extend(params_data)
            summary_data.append([])  # 空行分隔

            # 添加原坐标系数据到汇总表
            summary_data.append(["原坐标"])
            summary_data.append(["序号", "台站名称", "纬度", "经度", "高程", "ECEF X", "ECEF Y", "ECEF Z"])
            summary_data.extend(coords1_data)
            summary_data.append([])  # 空行分隔

            # 添加新坐标系数据到汇总表
            summary_data.append(["新坐标"])
            summary_data.append(["序号", "台站名称", "纬度", "经度", "高程", "ECEF X", "ECEF Y", "ECEF Z"])
            summary_data.extend(coords2_data)
            summary_data.append([])  # 空行分隔

            # 添加差值数据到汇总表
            summary_data.append(["残差"])
            summary_data.append(["序号", "台站名称", "纬度残差", "经度残差", "高程残差", "X残差", "Y残差", "Z残差"])
            summary_data.extend(diff_data)
            summary_data.append([])  # 空行分隔

            summary_data.append(["残差最值"])
            summary_data.append(["最大纬度残差", "最小纬度残差", "最大经度残差", "最小经度残差", "最大高程残差", "最小高程残差", "最大X残差", "最小X残差", "最大Y残差", "最小Y残差", "最大Z残差", "最小Z残差"])
            summary_data.extend(diff_sort_data)
            summary_data.append([])  # 空行分隔

            # 添加 Mp, Mx, My, Mz 到汇总表
            summary_data.append(["点位中误差"])
            summary_data.append(["Mx", "My", "Mz", "Mp","3Mx", "3My", "3Mz", "3Mp"])
            summary_data.extend(residual_summary_data)
            summary_data.append([])  # 空行分隔

            # 添加“读取坐标系1的文件”数据到汇总表
            summary_data.append(["转换前坐标"])
            summary_data.append(["序号", "台站名称", "纬度", "经度", "高程", "ECEF X", "ECEF Y", "ECEF Z"])
            summary_data.extend(coords_to_transform_data)
            summary_data.append([])  # 空行分隔

            # 添加转换后坐标数据到汇总表
            summary_data.append(["转换后坐标"])
            summary_data.append(["序号", "台站名称", "纬度", "经度", "高程", "ECEF X", "ECEF Y", "ECEF Z"])
            summary_data.extend(transformed_data)
            summary_data.append([])  # 空行分隔

            summary_data.append(["改正数"])
            summary_data.append(["序号", "台站名称", "Vx", "Vy", "Vz"])
            summary_data.extend(correction_data)
            summary_data.append([])  # 空行分隔

            summary_data.append(["改正后坐标"])
            summary_data.append(["序号", "台站名称", "纬度(改)", "经度(改)", "高程(改)", "X(改)", "Y(改)", "Z(改)"])
            summary_data.extend(corrected_data)


            # 写入汇总表
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name="汇总", index=False, header=False)

            # 写入单独的工作表
            for sheet_name, df in data_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        # 自动调整列宽并设置单元格居中对齐
        workbook = load_workbook(file_path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for column_cells in sheet.columns:
                max_length = 0
                column = column_cells[0].column_letter  # 获取列字母
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                            cell.alignment = Alignment(horizontal="center", vertical="center")  # 设置居中对齐
                    except:
                        pass
                adjusted_width = (max_length + 2)  # 加2留出额外的空白
                sheet.column_dimensions[column].width = adjusted_width

            # 获取 3Mx、3My 和 3Mz 的值
            mx3, my3, mz3 = 0, 0, 0
            if self.residual_summary_tree.get_children():
                mx3, my3, mz3 = (
                    float(self.residual_summary_tree.item(self.residual_summary_tree.get_children()[0])["values"][4]),
                    float(self.residual_summary_tree.item(self.residual_summary_tree.get_children()[0])["values"][5]),
                    float(self.residual_summary_tree.item(self.residual_summary_tree.get_children()[0])["values"][6])
                )


        workbook.save(file_path)
        messagebox.showinfo("导出成功", f"数据成功导出到 {file_path}")

    def __init__(self, root):
        self.root = root
        self.root.title("坐标转换工具")

        # 设置窗口初始大小
        self.root.geometry("1600x900")
        # 允许窗口横向缩放
        self.root.columnconfigure(0, weight=1)
        self.root.columnconfigure(1, weight=1)
        self.root.columnconfigure(2, weight=1)
        self.root.columnconfigure(3, weight=1)

        # 允许窗口纵向扩展
        self.root.rowconfigure(0, weight=0)
        self.root.rowconfigure(1, weight=0)
        self.root.rowconfigure(2, weight=1)
        self.root.rowconfigure(3, weight=1)
        self.root.rowconfigure(4, weight=0)
        self.root.rowconfigure(5, weight=0)
        self.root.rowconfigure(6, weight=0)
        self.root.rowconfigure(7, weight=0)
        self.root.rowconfigure(8, weight=1)
        self.root.rowconfigure(10, weight=1)
        self.root.rowconfigure(9, weight=0)

        # 创建Treeview样式
        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Arial", 9, "bold"))  # 调整表头字体为9号
        style.configure("Treeview", rowheight=20, font=("Arial", 9))  # 调整行高和内容字体为9号

        # 第1行：长半径 和 扁率输入框
        tk.Label(root, text="长半径 (默认为 WGS84: 6378137.0 米):").grid(row=0, column=0, sticky='e', padx=10, pady=10)
        self.a_entry = tk.Entry(root)
        self.a_entry.grid(row=0, column=1, padx=10, pady=10, sticky='ew')
        self.a_entry.insert(0, "6378137.0")

        tk.Label(root, text="扁率 (默认为 WGS84: 0.003352810664747480):").grid(row=0, column=2, sticky='e', padx=10,
                                                                               pady=10)
        self.f_entry = tk.Entry(root)
        self.f_entry.grid(row=0, column=3, padx=10, pady=10, sticky='ew')
        self.f_entry.insert(0, "0.003352810664747480")

        # 第2行：导入原坐标 和 导入新坐标按钮
        self.file_button_1 = tk.Button(root, text="导入原坐标", command=self.load_file_1)
        self.file_button_1.grid(row=1, column=0, padx=10, pady=10)

        self.coord_type_var_1 = tk.StringVar()
        self.coord_type_combo_1 = ttk.Combobox(root, textvariable=self.coord_type_var_1, state="readonly")
        self.coord_type_combo_1['values'] = ("大地坐标", "空间直角坐标")
        self.coord_type_combo_1.grid(row=1, column=1, padx=10, pady=10)
        self.coord_type_combo_1.current(0)

        self.file_button_2 = tk.Button(root, text="导入新坐标", command=self.load_file_2)
        self.file_button_2.grid(row=1, column=2, padx=10, pady=10)

        self.coord_type_var_2 = tk.StringVar()
        self.coord_type_combo_2 = ttk.Combobox(root, textvariable=self.coord_type_var_2, state="readonly")
        self.coord_type_combo_2['values'] = ("大地坐标", "空间直角坐标")
        self.coord_type_combo_2.grid(row=1, column=3, padx=10, pady=10)
        self.coord_type_combo_2.current(0)

        # 第3行：原坐标和新坐标表格
        self.tree1 = ttk.Treeview(root, columns=("num", "name", "lat", "lon", "h", "x", "y", "z"), show="headings",
                                  height=10)
        self.tree1.grid(row=2, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")
        self.tree1.heading("num", text="序号")
        self.tree1.heading("name", text="台站名称")
        self.tree1.heading("lat", text="纬度")
        self.tree1.heading("lon", text="经度")
        self.tree1.heading("h", text="高程")
        self.tree1.heading("x", text="ECEF X")
        self.tree1.heading("y", text="ECEF Y")
        self.tree1.heading("z", text="ECEF Z")
        self.tree1.column("num", width=50, anchor='center')
        self.tree1.column("name", width=100, anchor='center')
        self.tree1.column("lat", width=80, anchor='center')
        self.tree1.column("lon", width=80, anchor='center')
        self.tree1.column("h", width=80, anchor='center')
        self.tree1.column("x", width=80, anchor='center')
        self.tree1.column("y", width=80, anchor='center')
        self.tree1.column("z", width=80, anchor='center')

        self.tree2 = ttk.Treeview(root, columns=("num", "name", "lat", "lon", "h", "x", "y", "z"), show="headings",
                                  height=10)
        self.tree2.grid(row=2, column=2,columnspan=2, padx=10, pady=10, sticky="nsew")
        self.tree2.heading("num", text="序号")
        self.tree2.heading("name", text="台站名称")
        self.tree2.heading("lat", text="纬度")
        self.tree2.heading("lon", text="经度")
        self.tree2.heading("h", text="高程")
        self.tree2.heading("x", text="ECEF X")
        self.tree2.heading("y", text="ECEF Y")
        self.tree2.heading("z", text="ECEF Z")
        self.tree2.column("num", width=50, anchor='center')
        self.tree2.column("name", width=100, anchor='center')
        self.tree2.column("lat", width=80, anchor='center')
        self.tree2.column("lon", width=80, anchor='center')
        self.tree2.column("h", width=80, anchor='center')
        self.tree2.column("x", width=80, anchor='center')
        self.tree2.column("y", width=80, anchor='center')
        self.tree2.column("z", width=80, anchor='center')

        # 第4行：残差表格，横跨两列
        # 修改残差表格，增加最大值和最小值的列
        self.diff_tree = ttk.Treeview(root, columns=(
            "num", "name", "lat_diff", "lon_diff", "h_diff", "x_diff", "y_diff", "z_diff"),
                                      show="headings", height=10)

        self.diff_tree.grid(row=3, column=0, columnspan=4, padx=10, pady=10, sticky="nsew")

        # 现有列
        self.diff_tree.heading("num", text="序号")
        self.diff_tree.heading("name", text="台站名称")
        self.diff_tree.heading("lat_diff", text="纬度残差")
        self.diff_tree.heading("lon_diff", text="经度残差")
        self.diff_tree.heading("h_diff", text="高程残差")
        self.diff_tree.heading("x_diff", text="X残差")
        self.diff_tree.heading("y_diff", text="Y残差")
        self.diff_tree.heading("z_diff", text="Z残差")


        self.diff_tree.column("num", width=50, anchor='center')
        self.diff_tree.column("name", width=100, anchor='center')
        self.diff_tree.column("lat_diff", width=80, anchor='center')
        self.diff_tree.column("lon_diff", width=80, anchor='center')
        self.diff_tree.column("h_diff", width=80, anchor='center')
        self.diff_tree.column("x_diff", width=80, anchor='center')
        self.diff_tree.column("y_diff", width=80, anchor='center')
        self.diff_tree.column("z_diff", width=80, anchor='center')



        self.diff_sort_tree = ttk.Treeview(root, columns=("lat_max", "lat_min", "lon_max", "lon_min", "h_max", "h_min", "x_max", "x_min", "y_max", "y_min", "z_max",
            "z_min"), show="headings", height=1)
        self.diff_sort_tree.grid(row=4, column=0, columnspan=4, padx=10, pady=10, sticky="nsew")

        # 新增的最大值最小值列
        self.diff_sort_tree.heading("lat_max", text="最大纬度残差")
        self.diff_sort_tree.heading("lat_min", text="最小纬度残差")
        self.diff_sort_tree.heading("lon_max", text="最大经度残差")
        self.diff_sort_tree.heading("lon_min", text="最小经度残差")
        self.diff_sort_tree.heading("h_max", text="最大高程残差")
        self.diff_sort_tree.heading("h_min", text="最小高程残差")
        self.diff_sort_tree.heading("x_max", text="最大X残差")
        self.diff_sort_tree.heading("x_min", text="最小X残差")
        self.diff_sort_tree.heading("y_max", text="最大Y残差")
        self.diff_sort_tree.heading("y_min", text="最小Y残差")
        self.diff_sort_tree.heading("z_max", text="最大Z残差")
        self.diff_sort_tree.heading("z_min", text="最小Z残差")

        self.diff_sort_tree.column("lat_max", width=80, anchor='center')
        self.diff_sort_tree.column("lat_min", width=80, anchor='center')
        self.diff_sort_tree.column("lon_max", width=80, anchor='center')
        self.diff_sort_tree.column("lon_min", width=80, anchor='center')
        self.diff_sort_tree.column("h_max", width=80, anchor='center')
        self.diff_sort_tree.column("h_min", width=80, anchor='center')
        self.diff_sort_tree.column("x_max", width=80, anchor='center')
        self.diff_sort_tree.column("x_min", width=80, anchor='center')
        self.diff_sort_tree.column("y_max", width=80, anchor='center')
        self.diff_sort_tree.column("y_min", width=80, anchor='center')
        self.diff_sort_tree.column("z_max", width=80, anchor='center')
        self.diff_sort_tree.column("z_min", width=80, anchor='center')



        # 在第4行初始化残差表格下方添加新的表格
        self.residual_summary_tree = ttk.Treeview(root, columns=("Mx", "My", "Mz", "Mp","3Mx", "3My", "3Mz", "3Mp"), show="headings", height=1)
        self.residual_summary_tree.grid(row=5, column=0, columnspan=4, padx=10, pady=10, sticky="nsew")
        self.residual_summary_tree.heading("Mx", text="Mx")
        self.residual_summary_tree.heading("My", text="My")
        self.residual_summary_tree.heading("Mz", text="Mz")
        self.residual_summary_tree.heading("Mp", text="Mp")
        self.residual_summary_tree.heading("3Mx", text="3Mx")
        self.residual_summary_tree.heading("3My", text="3My")
        self.residual_summary_tree.heading("3Mz", text="3Mz")
        self.residual_summary_tree.heading("3Mp", text="3Mp")
        self.residual_summary_tree.column("Mx", width=80, anchor='center')
        self.residual_summary_tree.column("My", width=80, anchor='center')
        self.residual_summary_tree.column("Mz", width=80, anchor='center')
        self.residual_summary_tree.column("Mp", width=80, anchor='center')
        self.residual_summary_tree.column("3Mx", width=80, anchor='center')
        self.residual_summary_tree.column("3My", width=80, anchor='center')
        self.residual_summary_tree.column("3Mz", width=80, anchor='center')
        self.residual_summary_tree.column("3Mp", width=80, anchor='center')

        # 第5行：计算七参数 和 七参数表格
        self.compute_button = tk.Button(root, text="计算七参数", command=self.compute_seven_parameters)
        self.compute_button.grid(row=6, column=0, padx=10, pady=10)

        self.param_tree = ttk.Treeview(root, columns=("dx", "dy", "dz", "rx", "ry", "rz", "m", "xx"), show="headings",
                                       height=1)
        self.param_tree.grid(row=6, column=1, columnspan=3, padx=10, pady=10, sticky="nsew")
        self.param_tree.heading("dx", text="dx(m)")
        self.param_tree.heading("dy", text="dy(m)")
        self.param_tree.heading("dz", text="dz(m)")
        self.param_tree.heading("rx", text="rx(s)")
        self.param_tree.heading("ry", text="ry(s)")
        self.param_tree.heading("rz", text="rz(s)")
        self.param_tree.heading("m", text="m(ppm)")
        self.param_tree.heading("xx", text="单位权中误差")
        self.param_tree.column("dx", width=100, anchor='center')
        self.param_tree.column("dy", width=100, anchor='center')
        self.param_tree.column("dz", width=100, anchor='center')
        self.param_tree.column("rx", width=100, anchor='center')
        self.param_tree.column("ry", width=100, anchor='center')
        self.param_tree.column("rz", width=100, anchor='center')
        self.param_tree.column("m", width=100, anchor='center')
        self.param_tree.column("xx", width=100, anchor='center')

        # 第6行：导入需要转换的坐标 和 使用七参数转换
        self.file_button_3 = tk.Button(root, text="导入需要转换的坐标", command=self.load_file_to_transform)
        self.file_button_3.grid(row=7, column=0, padx=10, pady=10)

        self.coord_type_var_3 = tk.StringVar()
        self.coord_type_combo_3 = ttk.Combobox(root, textvariable=self.coord_type_var_3, state="readonly")
        self.coord_type_combo_3['values'] = ("大地坐标", "空间直角坐标")
        self.coord_type_combo_3.grid(row=7, column=1, padx=10, pady=10)
        self.coord_type_combo_3.current(0)

        self.compute_button_2 = tk.Button(root, text="使用七参数转换", command=self.transform_coords)
        self.compute_button_2.grid(row=7, column=2, columnspan=2, padx=10, pady=10)

        # 第7行：转换前后表格
        self.transform_input_tree = ttk.Treeview(root, columns=("num", "name", "lat", "lon", "h", "x", "y", "z"),
                                                 show="headings", height=10)
        self.transform_input_tree.grid(row=8, column=0, columnspan=2,padx=10, pady=10, sticky="nsew")
        self.transform_input_tree.heading("num", text="序号")
        self.transform_input_tree.heading("name", text="台站名称")
        self.transform_input_tree.heading("lat", text="纬度")
        self.transform_input_tree.heading("lon", text="经度")
        self.transform_input_tree.heading("h", text="高程")
        self.transform_input_tree.heading("x", text="ECEF X")
        self.transform_input_tree.heading("y", text="ECEF Y")
        self.transform_input_tree.heading("z", text="ECEF Z")
        self.transform_input_tree.column("num", width=50, anchor='center')
        self.transform_input_tree.column("name", width=100, anchor='center')
        self.transform_input_tree.column("lat", width=80, anchor='center')
        self.transform_input_tree.column("lon", width=80, anchor='center')
        self.transform_input_tree.column("h", width=80, anchor='center')
        self.transform_input_tree.column("x", width=80, anchor='center')
        self.transform_input_tree.column("y", width=80, anchor='center')
        self.transform_input_tree.column("z", width=80, anchor='center')

        self.transformed_output_tree = ttk.Treeview(root, columns=("num", "name", "lat", "lon", "h", "x", "y", "z"),
                                                    show="headings", height=10)
        self.transformed_output_tree.grid(row=8, column=2, columnspan=2, padx=10, pady=10, sticky="nsew")
        self.transformed_output_tree.heading("num", text="序号")
        self.transformed_output_tree.heading("name", text="台站名称")
        self.transformed_output_tree.heading("lat", text="纬度")
        self.transformed_output_tree.heading("lon", text="经度")
        self.transformed_output_tree.heading("h", text="高程")
        self.transformed_output_tree.heading("x", text="ECEF X")
        self.transformed_output_tree.heading("y", text="ECEF Y")
        self.transformed_output_tree.heading("z", text="ECEF Z")
        self.transformed_output_tree.column("num", width=50, anchor='center')
        self.transformed_output_tree.column("name", width=100, anchor='center')
        self.transformed_output_tree.column("lat", width=80, anchor='center')
        self.transformed_output_tree.column("lon", width=80, anchor='center')
        self.transformed_output_tree.column("h", width=80, anchor='center')
        self.transformed_output_tree.column("x", width=80, anchor='center')
        self.transformed_output_tree.column("y", width=80, anchor='center')
        self.transformed_output_tree.column("z", width=80, anchor='center')

        # 新增第9行：改正数表格
        self.correction_tree = ttk.Treeview(root, columns=("num", "name", "Vx", "Vy", "Vz"), show="headings", height=10)
        self.correction_tree.grid(row=10, column=0, columnspan=2, padx=10, pady=10, sticky="nsew")
        self.correction_tree.heading("num", text="序号")
        self.correction_tree.heading("name", text="台站名称")
        self.correction_tree.heading("Vx", text="Vx")
        self.correction_tree.heading("Vy", text="Vy")
        self.correction_tree.heading("Vz", text="Vz")

        self.correction_tree.column("num", width=50, anchor='center')
        self.correction_tree.column("name", width=100, anchor='center')
        self.correction_tree.column("Vx", width=80, anchor='center')
        self.correction_tree.column("Vy", width=80, anchor='center')
        self.correction_tree.column("Vz", width=80, anchor='center')

        # 新增第10行：改正后结果表格
        self.corrected_tree = ttk.Treeview(root, columns=(
        "num", "name", "lat_corrected", "lon_corrected", "h_corrected", "X_corrected", "Y_corrected", "Z_corrected"),
                                           show="headings", height=10)
        self.corrected_tree.grid(row=10, column=2, columnspan=2, padx=10, pady=10, sticky="nsew")
        self.corrected_tree.heading("num", text="序号")
        self.corrected_tree.heading("name", text="台站名称")
        self.corrected_tree.heading("lat_corrected", text="纬度(改)")
        self.corrected_tree.heading("lon_corrected", text="经度(改)")
        self.corrected_tree.heading("h_corrected", text="高程(改)")
        self.corrected_tree.heading("X_corrected", text="X(改)")
        self.corrected_tree.heading("Y_corrected", text="Y(改)")
        self.corrected_tree.heading("Z_corrected", text="Z(改)")

        self.corrected_tree.column("num", width=50, anchor='center')
        self.corrected_tree.column("name", width=100, anchor='center')
        self.corrected_tree.column("lat_corrected", width=100, anchor='center')
        self.corrected_tree.column("lon_corrected", width=100, anchor='center')
        self.corrected_tree.column("h_corrected", width=100, anchor='center')
        self.corrected_tree.column("X_corrected", width=80, anchor='center')
        self.corrected_tree.column("Y_corrected", width=80, anchor='center')
        self.corrected_tree.column("Z_corrected", width=80, anchor='center')

        # 修改计算改正数按钮的 command
        self.compute_button_x = tk.Button(root, text="计算改正数", command=self.compute_correction)
        self.compute_button_x.grid(row=9, column=0, columnspan=2, padx=10, pady=10)

        # 修改使用改正数平差按钮的 command
        self.compute_button_y = tk.Button(root, text="使用改正数平差", command=self.apply_correction_adjustment)
        self.compute_button_y.grid(row=9, column=2, columnspan=2, padx=10, pady=10)

        # 第8行：导出为Excel按钮，居中
        self.export_button = tk.Button(root, text="导出为Excel", command=self.export_to_excel)
        self.export_button.grid(row=11, column=0, columnspan=4, padx=10, pady=20, sticky="ew")

        # 坐标系数据初始化
        self.coords1 = []
        self.coords2 = []
        self.coords_to_transform = []

    def check_and_calculate_correction(self):
        """
        检查是否已导入所有必要文件并计算非公共点改正数
        """
        # 确保公共点1、公共点2和非公共点数据都已导入
        if self.coords1 and self.coords2 and self.coords_to_transform:
            self.compute_and_update_non_public_correction()

    def load_file_1(self):
        file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
        if file_path:
            stations = read_file(file_path)
            if not stations:
                messagebox.showerror("错误", "文件内容无效或空")
                return

            # 清空之前的数据
            for row in self.tree1.get_children():
                self.tree1.delete(row)
            self.coords1 = []  # 重置原坐标列表
            self.params = None  # 重置七参数

            # 判断选择的数据类型
            coord_type = self.coord_type_var_1.get()

            # 获取用户输入的椭球参数
            a = float(self.a_entry.get())
            f = float(self.f_entry.get())

            for station in stations:
                if coord_type == "大地坐标":
                    lat = station['lat']
                    lon = station['lon']
                    h = station['h']
                    x, y, z = lat_lon_to_ecef(lat, lon, h, a, f)
                    self.coords1.append((x, y, z))
                    self.tree1.insert("", "end", values=(
                    station['num'], station['name'], f"{lat:.8f}", f"{lon:.8f}", f"{h:.8f}", f"{x:.8f}", f"{y:.8f}",
                    f"{z:.8f}"))
                else:
                    x, y, z = station['lon'], station['lat'], station['h']
                    lat, lon, h = ecef_to_lat_lon_ecef(x, y, z)
                    self.coords1.append((x, y, z))
                    self.tree1.insert("", "end", values=(
                    station['num'], station['name'], f"{lat:.8f}", f"{lon:.8f}", f"{h:.8f}", f"{x:.8f}", f"{y:.8f}",
                    f"{z:.8f}"))

            if self.coords2:
                self.compute_and_update_diff()

    def load_file_2(self):
        file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
        if file_path:
            stations = read_file(file_path)
            if not stations:
                messagebox.showerror("错误", "文件内容无效或空")
                return

            # 清空之前的数据
            for row in self.tree2.get_children():
                self.tree2.delete(row)
            self.coords2 = []  # 重置新坐标列表
            self.params = None  # 重置七参数

            coord_type = self.coord_type_var_2.get()
            a = float(self.a_entry.get())
            f = float(self.f_entry.get())

            for station in stations:
                if coord_type == "大地坐标":
                    lat = station['lat']
                    lon = station['lon']
                    h = station['h']
                    x, y, z = lat_lon_to_ecef(lat, lon, h, a, f)
                    self.coords2.append((x, y, z))
                    self.tree2.insert("", "end", values=(
                    station['num'], station['name'], f"{lat:.8f}", f"{lon:.8f}", f"{h:.8f}", f"{x:.8f}", f"{y:.8f}",
                    f"{z:.8f}"))
                else:
                    x, y, z = station['lon'], station['lat'], station['h']
                    lat, lon, h = ecef_to_lat_lon_ecef(x, y, z)
                    self.coords2.append((x, y, z))
                    self.tree2.insert("", "end", values=(
                    station['num'], station['name'], f"{lat:.8f}", f"{lon:.8f}", f"{h:.8f}", f"{x:.8f}", f"{y:.8f}",
                    f"{z:.8f}"))

            if self.coords1:
                self.compute_and_update_diff()

    def compute_and_update_diff(self):
        """
        计算差值并更新到差值表格，包括纬度、经度、高程和ECEF的差值
        """
        # 清空差值表格
        for row in self.diff_tree.get_children():
            self.diff_tree.delete(row)

        # 确保七参数已经计算
        if not hasattr(self, 'params') or self.params is None:
            if not self.coords1 or not self.coords2:
                messagebox.showerror("错误", "请先加载两个坐标系的文件")
                return

            if len(self.coords1) != len(self.coords2):
                messagebox.showerror("错误", "两个坐标系的点数量不一致")
                return

            # 计算七参数并存储在self.params中
            self.params = compute_seven_parameters(self.coords1, self.coords2)

        # 进行坐标转换，使用计算出的self.params
        for row1, row2 in zip(self.tree1.get_children(), self.tree2.get_children()):
            # 获取 tree1 和 tree2 的每一行值
            values1 = self.tree1.item(row1)["values"]
            values2 = self.tree2.item(row2)["values"]

            # 提取ECEF坐标并进行七参数转换
            x1, y1, z1 = float(values1[5]), float(values1[6]), float(values1[7])
            transformed_coords = transform_to_second_coordinate_system([(x1, y1, z1)], self.params)[0]

            # 提取坐标系2的ECEF坐标
            x2, y2, z2 = float(values2[5]), float(values2[6]), float(values2[7])

            # 计算转换后的差值（X, Y, Z）
            x_diff = round(x2 - transformed_coords[0], 6)
            y_diff = round(y2 - transformed_coords[1], 6)
            z_diff = round(z2 - transformed_coords[2], 6)

            # 转换为纬度、经度和高程
            lat1, lon1, h1 = ecef_to_lat_lon_ecef(transformed_coords[0], transformed_coords[1], transformed_coords[2])
            lat2, lon2, h2 = float(values2[2]), float(values2[3]), float(values2[4])

            # 计算纬度、经度、高程的差值
            lat_diff = round(lat1 - lat2, 6)
            lon_diff = round(lon1 - lon2, 6)
            h_diff = round(h1 - h2, 6)

            # 插入差值到新的表格（包含纬度、经度、高程、X, Y, Z差值）
            self.diff_tree.insert("", "end", values=(
                values1[0], values1[1], f"{lat_diff:.8f}", f"{lon_diff:.8f}", f"{h_diff:.8f}", f"{x_diff:.8f}", f"{y_diff:.8f}", f"{z_diff:.8f}",
            ))
            

            # 清空残差表格
            for row in self.diff_sort_tree.get_children():
                self.diff_sort_tree.delete(row)

            lat_diffs, lon_diffs, h_diffs = [], [], []
            x_diffs, y_diffs, z_diffs = [], [], []

            for row1, row2 in zip(self.tree1.get_children(), self.tree2.get_children()):
                # coord_type1 = self.coord_type_var_1.get()
                # coord_type2 = self.coord_type_var_2.get()
                #
                # if coord_type1 == "大地坐标":
                #     a = float(self.a_entry.get())
                #     f = float(self.f_entry.get())
                #     values1 = self.tree1.item(row1)["values"]
                #     x, y, z = lat_lon_to_ecef(float(values1[2]), float(values1[3]), float(values1[4]), a, f)
                #     values2 = (values1[0], values1[1], values1[2], values1[3], values1[4], x, y, z)
                values1 = self.tree1.item(row1)["values"]
                values2 = self.tree2.item(row2)["values"]

                # 提取ECEF坐标并进行转换
                x1, y1, z1 = float(values1[5]), float(values1[6]), float(values1[7])
                x2, y2, z2 = float(values2[5]), float(values2[6]), float(values2[7])
                # 计算差值
                transformed_coords = transform_to_second_coordinate_system([(x1, y1, z1)], self.params)[0]
                # 计算转换后的差值（X, Y, Z）
                x_diff = round(x2 - transformed_coords[0], 6)
                y_diff = round(y2 - transformed_coords[1], 6)
                z_diff = round(z2 - transformed_coords[2], 6)

                # 转换为纬度、经度和高程
                # 转换为纬度、经度和高程
                lat1, lon1, h1 = ecef_to_lat_lon_ecef(transformed_coords[0], transformed_coords[1],
                                                      transformed_coords[2])
                lat2, lon2, h2 = float(values2[2]), float(values2[3]), float(values2[4])

                lat_diff = round(lat1 - lat2, 6)
                lon_diff = round(lon1 - lon2, 6)
                h_diff = round(h1 - h2, 6)

                # 记录每次的差值
                lat_diffs.append(lat_diff)
                lon_diffs.append(lon_diff)
                h_diffs.append(h_diff)
                x_diffs.append(x_diff)
                y_diffs.append(y_diff)
                z_diffs.append(z_diff)

            # 计算最大值和最小值
            # lat_max, lat_min = max(lat_diffs, key=abs), min(lat_diffs, key=abs)
            # lon_max, lon_min = max(lon_diffs, key=abs), min(lon_diffs, key=abs)
            # h_max, h_min = max(h_diffs, key=abs), min(h_diffs, key=abs)
            # x_max, x_min = max(x_diffs, key=abs), min(x_diffs, key=abs)
            # y_max, y_min = max(y_diffs, key=abs), min(y_diffs, key=abs)
            # z_max, z_min = max(z_diffs, key=abs), min(z_diffs, key=abs)
            lat_max, lat_min = f"{max(lat_diffs, key=abs):.8f}", f"{min(lat_diffs, key=abs):.8f}"
            lon_max, lon_min = f"{max(lon_diffs, key=abs):.8f}", f"{min(lon_diffs, key=abs):.8f}"
            h_max, h_min = f"{max(h_diffs, key=abs):.8f}", f"{min(h_diffs, key=abs):.8f}"
            x_max, x_min = f"{max(x_diffs, key=abs):.8f}", f"{min(x_diffs, key=abs):.8f}"
            y_max, y_min = f"{max(y_diffs, key=abs):.8f}", f"{min(y_diffs, key=abs):.8f}"
            z_max, z_min = f"{max(z_diffs, key=abs):.8f}", f"{min(z_diffs, key=abs):.8f}"


            # 插入最大值最小值到表格
            self.diff_sort_tree.insert("", "end", values=(
                lat_max, lat_min, lon_max, lon_min, h_max, h_min, x_max, x_min, y_max, y_min, z_max, z_min
            ))

            # 初始化残差平方和
            vv_x = vv_y = vv_z = 0
            n = len(self.tree1.get_children())

            for row1, row2 in zip(self.tree1.get_children(), self.tree2.get_children()):
                ...
                # 计算平方和
                vv_x += x_diff ** 2
                vv_y += y_diff ** 2
                vv_z += z_diff ** 2

            # 计算 Mx, My, Mz, Mp
            Mx = math.sqrt(vv_x / (n - 1))
            My = math.sqrt(vv_y / (n - 1))
            Mz = math.sqrt(vv_z / (n - 1))
            Mp = math.sqrt(Mx ** 2 + My ** 2 + Mz ** 2)
            Mx3 = 3 * Mx
            My3 = 3 * My
            Mz3 = 3 * Mz
            Mp3 = 3 * Mp
            # # 计算 Mx, My, Mz, Mp
            # Mx = math.sqrt(vv_x / (3*n - 7))
            # My = math.sqrt(vv_y / (3*n - 7))
            # Mz = math.sqrt(vv_z / (3*n - 7))
            # Mp = math.sqrt(Mx ** 2 + My ** 2 + Mz ** 2)
            # Mx3 = 3 * Mx
            # My3 = 3 * My
            # Mz3 = 3 * Mz
            # Mp3 = 3 * Mp

            # 插入结果到表格
            for row in self.residual_summary_tree.get_children():
                self.residual_summary_tree.delete(row)

            self.residual_summary_tree.insert("", "end", values=(f"{Mx:.8f}", f"{My:.8f}", f"{Mz:.8f}", f"{Mp:.8f}",f"{Mx3:.8f}", f"{My3:.8f}", f"{Mz3:.8f}", f"{Mp3:.8f}"))

        mx3, my3, mz3 = float(
            self.residual_summary_tree.item(self.residual_summary_tree.get_children()[0])["values"][4]), \
            float(self.residual_summary_tree.item(self.residual_summary_tree.get_children()[0])["values"][5]), \
            float(self.residual_summary_tree.item(self.residual_summary_tree.get_children()[0])["values"][6])

        # 检查差值是否超出3Mx、3My、3Mz，并标记界面上的超出值
        # 初始化标志变量，确保警告窗口只弹出一次
        warning_shown = False

        for row in self.diff_tree.get_children():
            values = self.diff_tree.item(row)["values"]
            x_diff, y_diff, z_diff = float(values[5]), float(values[6]), float(values[7])

            # 检查是否有异常值
            if abs(x_diff) > mx3 or abs(y_diff) > my3 or abs(z_diff) > mz3:
                # 如果没有弹出过警告窗口，则弹出
                if not warning_shown:
                    messagebox.showwarning("数据异常", "残差表中存在异常值")
                    warning_shown = True  # 设置标志为True，防止再次弹出窗口

                # 标记并高亮异常值
                for i, diff in enumerate([x_diff, y_diff, z_diff]):
                    if abs(diff) > [mx3, my3, mz3][i]:
                        self.diff_tree.tag_configure("highlight", background="red", foreground="white")
                        self.diff_tree.item(row, tags="highlight")



    def compute_seven_parameters(self):

        if not self.coords1 or not self.coords2:
            messagebox.showerror("错误", "请先加载两个坐标系的文件")
            return

        if len(self.coords1) != len(self.coords2):
            messagebox.showerror("错误", "两个坐标系的点数量不一致")
            return

        seven_params = compute_seven_parameters_sec(self.coords1, self.coords2)

        for row in self.param_tree.get_children():
            self.param_tree.delete(row)

        # 插入七参数结果
        self.param_tree.insert("", "end",
                               values=(f"{seven_params[0]:.10f}", f"{seven_params[1]:.10f}", f"{seven_params[2]:.10f}",
                                       f"{seven_params[3]:.10f}", f"{seven_params[4]:.10f}", f"{seven_params[5]:.10f}",
                                       f"{seven_params[6]:.10f}", f"{seven_params[7]:.10f}"))




    def load_file_to_transform(self):
        """
        加载需要转换的txt文件，读取经纬度和高程或XYZ，并显示在表格中。
        """
        file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])
        if file_path:
            stations = read_file(file_path)
            if not stations:
                messagebox.showerror("错误", "文件内容无效或空")
                return

            for row in self.transform_input_tree.get_children():
                self.transform_input_tree.delete(row)

            # 判断选择的数据类型
            coord_type = self.coord_type_var_3.get()

            # 获取用户输入的椭球参数
            a = float(self.a_entry.get())
            f = float(self.f_entry.get())

            self.coords_to_transform = []
            for station in stations:
                if coord_type == "大地坐标":
                    # 大地坐标转换为ECEF
                    lat = station['lat']
                    lon = station['lon']
                    h = station['h']
                    x, y, z = lat_lon_to_ecef(lat, lon, h, a, f)
                    self.coords_to_transform.append((station['num'], station['name'], x, y, z))
                    # 插入经纬度和计算的ECEF坐标
                    self.transform_input_tree.insert("", "end", values=(
                        station['num'], station['name'], f"{lat:.8f}", f"{lon:.8f}", f"{h:.8f}", f"{x:.8f}", f"{y:.8f}",
                        f"{z:.8f}"
                    ))
                else:
                    # 空间直角坐标直接使用
                    x, y, z = station['lon'], station['lat'], station['h']
                    lat, lon, h = ecef_to_lat_lon_ecef(x, y, z)  # ECEF 转为经纬度
                    self.coords_to_transform.append((station['num'], station['name'], x, y, z))
                    # 插入XYZ和计算的经纬度
                    self.transform_input_tree.insert("", "end", values=(
                        station['num'], station['name'], f"{lat:.8f}", f"{lon:.8f}", f"{h:.8f}", f"{x:.8f}", f"{y:.8f}",
                        f"{z:.8f}"
                    ))
        # self.check_and_calculate_correction()

    def transform_coords(self):
        """
        将加载的坐标通过七参数转换到新的坐标系，并将ECEF坐标转换为经纬度和高程。
        """

        if not self.coords_to_transform:
            messagebox.showerror("错误", "请先加载需要转换的文件")
            return

        if not self.param_tree.get_children():
            messagebox.showerror("错误", "请先计算七参数")
            return

        # 获取七参数
        params = compute_seven_parameters(self.coords1, self.coords2)

        # 转换坐标
        transformed_coords = transform_to_second_coordinate_system(
            [(x, y, z) for _, _, x, y, z in self.coords_to_transform], params
        )

        # 清除转换后的表格
        for row in self.transformed_output_tree.get_children():
            self.transformed_output_tree.delete(row)

        # 将转换后的坐标以及台站名称插入到表格中
        for (num, name, x_old, y_old, z_old), (x_new, y_new, z_new) in zip(self.coords_to_transform,
                                                                           transformed_coords):
            # ECEF 转为 经纬度和高程
            lat, lon, h = ecef_to_lat_lon_ecef(x_new, y_new, z_new)

            # 插入转换后的数据，包括台站名称
            self.transformed_output_tree.insert("", "end", values=(
                num, name, f"{lat:.8f}", f"{lon:.8f}", f"{h:.8f}", f"{x_new:.8f}", f"{y_new:.8f}", f"{z_new:.8f}"
            ))

    def compute_and_update_non_public_correction(self):
        """
        计算非公共点的改正数，并更新结果到表格中
        """
        if not self.coords1 or not self.coords2 or not self.coords_to_transform:
            messagebox.showerror("错误", "请先加载公共点和非公共点坐标文件")
            return

        # 确保七参数已经计算
        if not hasattr(self, 'params') or self.params is None:
            messagebox.showerror("错误", "请先计算七参数")
            return

        # 将坐标转换为numpy数组
        XYZ1 = np.array(self.coords1)
        XYZ2 = np.array(self.coords2)
        XYZ_non_public = np.array([(x, y, z) for _, _, x, y, z in self.coords_to_transform])

        # Step 1: 使用七参数转换非公共点坐标 (coords_to_transform)
        # transformed_non_public_points = self.transform_to_second_coordinate_system2(XYZ_non_public, self.params)
        # 获取七参数
        params = compute_seven_parameters(self.coords1, self.coords2)

        # 转换坐标
        transformed_non_public_points = transform_to_second_coordinate_system(
            [(x, y, z) for _, _, x, y, z in self.coords_to_transform], params
        )


        # Step 2: 计算公共点间的改正数
        corrections = compute_correction_with_seven_params(XYZ1, XYZ2, XYZ_non_public, self.params)

        # 清空现有的改正数表格数据
        for row in self.correction_tree.get_children():
            self.correction_tree.delete(row)

        # 清空改正后结果表格数据
        for row in self.corrected_tree.get_children():
            self.corrected_tree.delete(row)

        # Step 3: 处理计算结果，插入到表格中
        for i, ((num, name, x, y, z), transformed_point, correction) in enumerate(
                zip(self.coords_to_transform, transformed_non_public_points, corrections)):
            # 插入改正数表格
            self.correction_tree.insert("", "end", values=(
                num, name, f"{correction[0]:.8f}", f"{correction[1]:.8f}", f"{correction[2]:.8f}"
            ))

            # Step 4: 计算改正后的X、Y、Z (七参数转换后的值 + 改正数)
            X_corrected = transformed_point[0] + correction[0]
            Y_corrected = transformed_point[1] + correction[1]
            Z_corrected = transformed_point[2] + correction[2]

            # 将X_corrected, Y_corrected, Z_corrected转换回大地坐标系
            lat_corrected, lon_corrected, h_corrected = ecef_to_lat_lon_ecef(X_corrected, Y_corrected, Z_corrected)

            # 插入改正后结果表格
            self.corrected_tree.insert("", "end", values=(
                num, name, f"{lat_corrected:.8f}", f"{lon_corrected:.8f}", f"{h_corrected:.8f}", f"{X_corrected:.8f}",
                f"{Y_corrected:.8f}", f"{Z_corrected:.8f}"
            ))

    def compute_correction(self):
        if not self.coords1 or not self.coords2 or not self.coords_to_transform:
            messagebox.showerror("错误", "请先加载公共点和非公共点坐标文件")
            return

        self.params = compute_seven_parameters(self.coords1, self.coords2)

        XYZ1 = np.array(self.coords1)
        XYZ2 = np.array(self.coords2)
        XYZ_non_public = np.array([(x, y, z) for _, _, x, y, z in self.coords_to_transform])

        corrections = compute_correction_with_seven_params(XYZ1, XYZ2, XYZ_non_public, self.params)

        # 清空改正数表格
        for row in self.correction_tree.get_children():
            self.correction_tree.delete(row)

        for i, (num, name, correction) in enumerate(
                zip([s[0] for s in self.coords_to_transform], [s[1] for s in self.coords_to_transform], corrections)):
            self.correction_tree.insert("", "end", values=(
            num, name, f"{correction[0]:.8f}", f"{correction[1]:.8f}", f"{correction[2]:.8f}"))

    def apply_correction_adjustment(self):
        if not self.coords1 or not self.coords2 or not self.coords_to_transform:
            messagebox.showerror("错误", "请先加载公共点和非公共点坐标文件")
            return

        if self.params is None:
            self.params = compute_seven_parameters(self.coords1, self.coords2)

        transformed_non_public_points = transform_to_second_coordinate_system(
            [(x, y, z) for _, _, x, y, z in self.coords_to_transform], self.params
        )
        corrections = compute_correction_with_seven_params(
            np.array(self.coords1), np.array(self.coords2),
            np.array([(x, y, z) for _, _, x, y, z in self.coords_to_transform]), self.params
        )

        # 清空改正后结果表格
        for row in self.corrected_tree.get_children():
            self.corrected_tree.delete(row)

        for (num, name, _, _, _), transformed, correction in zip(self.coords_to_transform,
                                                                 transformed_non_public_points, corrections):
            X_corrected = transformed[0] + correction[0]
            Y_corrected = transformed[1] + correction[1]
            Z_corrected = transformed[2] + correction[2]
            lat_corrected, lon_corrected, h_corrected = ecef_to_lat_lon_ecef(X_corrected, Y_corrected, Z_corrected)
            self.corrected_tree.insert("", "end", values=(
            num, name, f"{lat_corrected:.8f}", f"{lon_corrected:.8f}", f"{h_corrected:.8f}", f"{X_corrected:.8f}",
            f"{Y_corrected:.8f}", f"{Z_corrected:.8f}"))


if __name__ == "__main__":
    root = tk.Tk()
    app = CoordinateConverterApp(root)
    root.mainloop()